import re
import pandas as pd
import streamlit as st
import glob
import os

# Detect all available squad-grid files
files = glob.glob("squad-grid-*.csv")
years = sorted([os.path.splitext(os.path.basename(f))[0].split("-")[-1] for f in files], reverse=True)

# Let the user choose the year
selected_year = st.selectbox("Select season year", years)

# Load the chosen CSV
file_path = f"squad-grid-{selected_year}.csv"

st.set_page_config(page_title=f"Oldham Athletic Squad Grid (season beginning {selected_year})", layout="wide")

# ---------- Load data ----------
df = pd.read_csv(file_path)

# ---------- Identify player columns (skip metadata/referee/etc.) ----------
META_COLS_KNOWN = {
    "unnamed: 0","date","opposition","goals1","goals2","venue","kickoff",
    "attendance","awayatt","post.position","opp.post.position","referee",
    "result","notes","competition","round"
}
def get_player_cols(df_: pd.DataFrame):
    meta_present = {c for c in df_.columns if c.lower() in META_COLS_KNOWN}
    with_space = [c for c in df_.columns if (" " in c and c not in meta_present)]
    return with_space if with_space else [c for c in df_.columns if c not in meta_present]

player_cols = get_player_cols(df)

# ---------- Event patterns ----------
EVENT_PATTERNS = [
    (r'\bog\s*(\d+)', lambda m: f'🔴⚽ {m.group(1)}'),   # own goal
    (r'\bpen\s*(\d+)', lambda m: f'🟢⚽ {m.group(1)}'),    # penalty goal
    (r'\bg\s*(\d+)', lambda m: f'⚽ {m.group(1)}'),      # normal goal
    (r'\by\s*(\d+)?', lambda m: f'🟨 {m.group(1)}' if m.group(1) else "🟨"), # yellow
    (r'\br\s*(\d+)?', lambda m: f'🟥 {m.group(1)}' if m.group(1) else "🟥"), # red
    (r'\bsub\s*\d*\s*on\s*(\d+)', lambda m: f'🔺 {m.group(1)}'),  # sub on
    (r'(\d+)\s*off', lambda m: f'🔻 {m.group(1)}'),                # sub off
    (r'\buu\b', lambda m: '🚫'),                                   # unused
    (r'\bx\b', lambda m: '🟩'),                                    # start
]

# ---------- Base background colours for role/status ----------
BG_START  = "#DFF0D8"  # light green
BG_SUB_ON = "#D9EDF7"  # light blue
BG_SUB_OFF= "#FCF8E3"  # pale yellow
BG_UNUSED = "#E6E6E6"  # light grey

def format_cell(raw) -> tuple[str, str]:
    """Return (display_text, background_colour) for a player's cell."""
    if raw is None:
        return "", ""
    s = str(raw).strip().lower()
    if not s or s == "nan":
        return "", ""

    # Determine background
    unused   = bool(re.search(r'\buu\b', s))
    sub_on_m = re.search(r'\bsub\s*\d*\s*on\s*(\d+)', s)
    off_m    = re.search(r'(\d+)\s*off', s)
    started  = bool(re.search(r'\bx\b', s))

    if unused:
        bg = BG_UNUSED
    elif sub_on_m:
        bg = BG_SUB_ON
    elif off_m:
        bg = BG_SUB_OFF
    elif started:
        bg = BG_START
    else:
        bg = ""

    # Build events in original order
    parts = []
    idx = 0
    while idx < len(s):
        match_found = False
        for pattern, repl in EVENT_PATTERNS:
            m = re.match(pattern, s[idx:], flags=re.IGNORECASE)
            if m:
                parts.append(repl(m))
                idx += m.end()
                match_found = True
                break
        if not match_found:
            idx += 1

    # Unused overrides display entirely
    if unused:
        parts = ["🚫"]

    return " ".join(parts).strip(), bg

# ---------- Build display DF + background map ----------
df_display = df.copy()
bg_map: dict[tuple[int, str], str] = {}

for col in player_cols:
    for i, val in df[col].items():
        disp, bg = format_cell(val)
        df_display.at[i, col] = disp
        if bg:
            bg_map[(i, col)] = bg

df_display = df_display.astype(str)

# ---------- Style for Streamlit ----------
def col_bg_styler(col: pd.Series):
    return [f"background-color: {bg_map.get((i, col.name), '')}" for i in col.index]

styled = df_display.style.apply(col_bg_styler, axis=0, subset=player_cols)

st.title("Oldham Athletic — Season Grid (emojis + role backgrounds)")
st.dataframe(styled, use_container_width=True)

# ---------- Legend ----------
legend_items = [
    "🟩 Start",
    "🔺 Sub on (minute)",
    "🔻 Sub off (minute)",
    "⚽ Goal (minute)",
    "🟢⚽ Penalty Goal (minute)",
    "🔴⚽ Own Goal (minute)",
    "🟨 Yellow Card (minute)",
    "🟥 Red Card (minute)",
    "🚫 Unused Sub"
]
legend_text = "\n".join(legend_items)
st.markdown("### Legend")
st.markdown("```\n" + legend_text + "\n```")

# ---------- Excel export ----------
def export_excel_with_bg_and_legend(df_disp: pd.DataFrame, bg_lookup: dict, filename: str = "squad_grid.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    wb = Workbook()
    ws = wb.active

    # headers
    for j, col in enumerate(df_disp.columns, start=1):
        ws.cell(row=1, column=j, value=col)

    # body
    for i in range(len(df_disp)):
        for j, col in enumerate(df_disp.columns, start=1):
            val = df_disp.iloc[i, j-1]
            cell = ws.cell(row=i+2, column=j, value=val)
            bg = bg_lookup.get((i, col))
            if bg:
                hex6 = bg.replace("#", "")
                cell.fill = PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")

    # legend after table
    legend_start_row = len(df_disp) + 4
    ws.cell(row=legend_start_row, column=1, value="Legend:")
    for idx, item in enumerate(legend_items, start=legend_start_row + 1):
        ws.cell(row=idx, column=1, value=item)

    return wb

if st.button("📥 Export to Excel"):
    from io import BytesIO
    bio = BytesIO()
    wb = export_excel_with_bg_and_legend(df_display, bg_map)
    wb.save(bio)
    st.download_button(
        "Download .xlsx",
        data=bio.getvalue(),
        file_name="squad_grid.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
